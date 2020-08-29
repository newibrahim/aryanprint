import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']


for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    correct_price = cell.value * 0.9
    corrected_cell = sheet.cell(row, 4)
    corrected_cell.value = correct_price

values = Reference(sheet,
                    min_row=2,
                    max_row=sheet.max_row,
                    min_col=4,
                    max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')
chart.type = 'bar'
new_cell = sheet.cell(5, 1)
price = 1043
new_cell.value = price

wb.save('transactions8.xlsx')


     





